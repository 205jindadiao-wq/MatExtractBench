# -*- coding: utf-8 -*-
import re
import json
import pandas as pd
from openpyxl import load_workbook
from pathlib import Path
from openai import OpenAI
from openpyxl.workbook import Workbook
from tqdm import tqdm
from typing import Dict, Any, List
import os

# ================= 1. 模型与 API 配置 =================
# 建议在服务器中先设置：
# export OPENAI_API_KEY="你的API_KEY"
CLIENT = OpenAI(
    api_key=os.getenv("OPENAI_API_KEY", "API_KEY"),
    base_url=""
)

MODEL_NAME = ""

# ================= 路径配置 =================
# 这里按纳米酶任务重新命名，你可以根据自己的真实路径修改
INPUT_EXCEL = "/data/nanozyme/纳米酶_eng.xlsx"
OUTPUT_EXCEL = "/data/nanozyme/nanozyme_result.xlsx"
CACHE_DIR = Path("/data/nanozyme/ocr_results_output_api")

# ================= 样品命名配置 =================
DEFAULT_NAME_PREFIX = "Nanozyme"
USE_DOI_IN_NAME = True

# ================= 2. 字段定义 =================
ELEMENT_COLUMNS = [
    "O", "N", "P", "S", "Si", "Se", "B", "F", "Cl", "Br", "I"
]

NANOZYME_PROPERTIES = [
    "Classification",
    "Metal oxidation state",
    "Shape",
    "Size (nm)",
    "Surface treatment",
    "Dispersion medium",
    "Optimal pH",
    "Optimal temperature (C)",
    "Substrate",
    "Km (mM)",
    "Vmax (uM s-1)",
    "Kcat (s-1)"
]

BASIC_COLUMNS = ["DOI", "Year", "Nanozyme Name"]

ALL_COLUMNS = (
    BASIC_COLUMNS
    + ["Classification"]
    + ELEMENT_COLUMNS
    + [
        "Metal oxidation state",
        "Shape",
        "Size (nm)",
        "Surface treatment",
        "Dispersion medium",
        "Optimal pH",
        "Optimal temperature (C)",
        "Substrate",
        "Km (mM)",
        "Vmax (uM s-1)",
        "Kcat (s-1)"
    ]
)

# ================= 3. 系统提示词 =================
SYSTEM_PROMPT = """你是一个专业的纳米酶数据挖掘系统，负责从纳米酶论文全文中提取结构化参数。

### 提取目标：
请针对文档中出现的每一种 nanozyme，提取以下参数，并严格输出 JSON。

### 需要提取的参数：

1. Nanozyme Name:
   - 纳米酶名称，例如 Fe3O4 NPs、CeO2 nanozyme、Pt nanoparticles、MoS2 nanosheets、carbon dots 等。
   - 如果文中有不同修饰、不同掺杂、不同复合结构，应分别作为不同 nanozyme 记录。

2. Classification:
   - 指明纳米酶属于哪一类酶活性，可多选，用分号分隔：
     - CAT: Catalase
     - HYL: Hydrolase
     - OXD: Oxidase
     - POD: Peroxidase
     - SOD: Superoxide Dismutase
   - 例如 "POD"、"OXD; POD"、"CAT; SOD"。
   - 如果没有明确酶类信息，填 "-"

3. Elemental composition:
   - 判断该 nanozyme 本身是否含有以下元素：
     O, N, P, S, Si, Se, B, F, Cl, Br, I
   - 存在填 "1"，不存在填 "0"，无法判断填 "-"
   - 注意：只判断 nanozyme 材料本身或明确接枝/修饰在 nanozyme 上的成分。
   - 不要把反应底物、缓冲液、测试试剂中的元素算入 nanozyme 组成。

4. Primary metal oxidation state:
   - 提取主要金属元素的氧化态，例如 Fe2+/Fe3+、Ce3+/Ce4+、Mn4+、Co2+/Co3+。
   - 如果是无金属纳米酶，例如 carbon dots、graphene oxide，填 "-"
   - 如果文中没有明确氧化态，不要猜测，填 "-"

5. Physical properties:
   - Shape: 形貌，例如 sphere、nanoparticle、nanosheet、nanorod、nanoflower、cube 等
   - Size (nm): 尺寸，单位 nm，可保留范围，如 "5-10"、"20 ± 3"
   - Surface treatment: 表面修饰或功能化，例如 PEG、PVP、citrate、aptamer、antibody、polymer coating
   - Dispersion medium: 分散介质，例如 water、PBS、buffer、ethanol 等

6. Enzymatic properties:
   - Optimal pH
   - Optimal temperature (C)
   - Substrate
   - Km (mM)
   - Vmax (uM s-1)
   - Kcat (s-1)

### 重要规则：

1. 严禁虚构：
   - 文档片段中没有出现的信息必须填 "-"
   - 不要根据常识猜测 Km、Vmax、Kcat、pH、temperature 等数值

2. 酶活性分类规则：
   - peroxidase-like activity → POD
   - oxidase-like activity → OXD
   - catalase-like activity → CAT
   - superoxide dismutase-like activity → SOD
   - hydrolase-like activity → HYL

3. 多底物 / 多酶活性：
   - 如果同一个 nanozyme 对不同 substrate 有不同 Km、Vmax、Kcat，应拆成多条记录
   - 如果同一个 nanozyme 同时具有多种酶活性，但没有分别给动力学参数，可以合并为一条记录，Classification 用分号连接

4. 单位处理：
   - Size 统一为 nm
   - Optimal temperature 只保留摄氏度数值
   - Km 统一为 mM
   - Vmax 统一为 uM s-1
   - Kcat 统一为 s-1
   - 如果原文单位不同但可以明确换算，可以换算；如果不能确定，保留原文数值和单位

5. JSON 输出格式必须严格如下：
[
  {
    "Nanozyme Name": "Fe3O4 nanoparticles",
    "Classification": "POD",
    "elements": {
      "O": "1",
      "N": "0",
      "P": "0",
      "S": "0",
      "Si": "0",
      "Se": "0",
      "B": "0",
      "F": "0",
      "Cl": "0",
      "Br": "0",
      "I": "0"
    },
    "properties": {
      "Metal oxidation state": "Fe2+/Fe3+",
      "Shape": "nanoparticle",
      "Size (nm)": "10",
      "Surface treatment": "-",
      "Dispersion medium": "PBS",
      "Optimal pH": "4.0",
      "Optimal temperature (C)": "37",
      "Substrate": "TMB",
      "Km (mM)": "0.098",
      "Vmax (uM s-1)": "3.44",
      "Kcat (s-1)": "-"
    }
  }
]
"""


# ================= 4. 辅助函数 =================
def split_text_into_chunks(text: str, chunk_size: int = 15000, overlap: int = 1500) -> List[str]:
    """将长文本切分为带重叠的片段"""
    chunks = []
    start = 0

    while start < len(text):
        end = start + chunk_size
        chunks.append(text[start:end])
        start += chunk_size - overlap

        if start >= len(text):
            break

    return chunks


def extract_nanozyme_candidates(text: str) -> List[str]:
    """
    从全文中粗略提取可能的纳米酶名称，用于给模型提示和样品名回退。
    该函数只做辅助，不作为最终结果。
    """
    candidates = []

    common_patterns = [
        r"\bFe3O4\b",
        r"\bFe2O3\b",
        r"\bCeO2\b",
        r"\bMnO2\b",
        r"\bCo3O4\b",
        r"\bCuO\b",
        r"\bCu2O\b",
        r"\bNiO\b",
        r"\bZnO\b",
        r"\bTiO2\b",
        r"\bV2O5\b",
        r"\bMoS2\b",
        r"\bWS2\b",
        r"\bPrussian blue\b",
        r"\bcarbon dots\b",
        r"\bgraphene oxide\b",
        r"\bAu NPs?\b",
        r"\bPt NPs?\b",
        r"\bPd NPs?\b",
        r"\bAg NPs?\b",
        r"\bAuNPs?\b",
        r"\bPtNPs?\b",
        r"\bPdNPs?\b",
        r"\bAgNPs?\b",
    ]

    for pat in common_patterns:
        candidates.extend(re.findall(pat, text, flags=re.IGNORECASE))

    # 捕获 “xxx nanozyme / xxx nanozymes”
    pattern_nanozyme = r"([A-Za-z0-9@\-/\+\.\(\)\[\] ]{2,80}?)\s+nanozymes?"
    matches = re.findall(pattern_nanozyme, text, flags=re.IGNORECASE)

    for m in matches:
        m = re.sub(r"\s+", " ", m).strip()
        if 2 <= len(m) <= 80:
            candidates.append(m)

    # 去重并保留顺序
    seen = set()
    clean_candidates = []

    for c in candidates:
        c = c.strip()
        key = c.lower()
        if key not in seen:
            seen.add(key)
            clean_candidates.append(c)

    return clean_candidates[:20]


def pick_first_valid(row_dict: Dict[str, Any], candidate_cols: List[str], default: str = "") -> str:
    """从多个可能列名中读取第一个有效值"""
    for col in candidate_cols:
        if col in row_dict:
            value = row_dict.get(col, "")
            if pd.notna(value) and str(value).strip() not in ["", "nan", "None"]:
                return str(value).strip()
    return default


def clean_json_content(content: str) -> str:
    """清理模型输出中的 Markdown 代码块"""
    content = content.strip()
    content = re.sub(r"```json", "", content, flags=re.IGNORECASE)
    content = re.sub(r"```", "", content)
    return content.strip()


def parse_llm_json(content: str):
    """从模型输出中解析 JSON"""
    content = clean_json_content(content)
    match = re.search(r"(\[.*\]|\{.*\})", content, re.DOTALL)

    if not match:
        return None

    parsed = json.loads(match.group(1))

    if isinstance(parsed, list):
        return parsed

    if isinstance(parsed, dict):
        return [parsed]

    return None


def normalize_element_value(value: Any) -> str:
    """规范元素存在性字段，只允许 1 / 0 / -"""
    if value is None:
        return "-"

    value = str(value).strip()

    if value in ["1", "yes", "Yes", "YES", "present", "Present", "true", "True"]:
        return "1"

    if value in ["0", "no", "No", "NO", "absent", "Absent", "false", "False"]:
        return "0"

    if value in ["-", "", "nan", "None", "unknown", "Unknown", "N/A", "NA"]:
        return "-"

    return value


def normalize_entry(entry: Dict[str, Any], target_name: str) -> Dict[str, Any]:
    """
    统一模型输出结构，防止模型偶尔输出字段不完整。
    返回扁平化后的字段字典。
    """
    elements = entry.get("elements", {})
    properties = entry.get("properties", {})

    # 兼容模型没有嵌套输出的情况
    flat = {}

    nanozyme_name = entry.get("Nanozyme Name", "-")
    if not nanozyme_name or str(nanozyme_name).strip() in ["", "-", "nan", "None"]:
        nanozyme_name = target_name

    flat["Nanozyme Name"] = str(nanozyme_name).strip()
    flat["Classification"] = str(entry.get("Classification", properties.get("Classification", "-"))).strip()

    for elem in ELEMENT_COLUMNS:
        flat[elem] = normalize_element_value(elements.get(elem, entry.get(elem, "-")))

    for prop in [
        "Metal oxidation state",
        "Shape",
        "Size (nm)",
        "Surface treatment",
        "Dispersion medium",
        "Optimal pH",
        "Optimal temperature (C)",
        "Substrate",
        "Km (mM)",
        "Vmax (uM s-1)",
        "Kcat (s-1)"
    ]:
        value = properties.get(prop, entry.get(prop, "-"))
        if value is None or str(value).strip() in ["", "nan", "None"]:
            value = "-"
        flat[prop] = str(value).strip()

    return flat


def is_useful_entry(entry: Dict[str, Any]) -> bool:
    """判断一条记录是否有有效信息"""
    check_cols = [
        "Nanozyme Name",
        "Classification",
        "Metal oxidation state",
        "Shape",
        "Size (nm)",
        "Surface treatment",
        "Dispersion medium",
        "Optimal pH",
        "Optimal temperature (C)",
        "Substrate",
        "Km (mM)",
        "Vmax (uM s-1)",
        "Kcat (s-1)"
    ]

    for col in check_cols:
        value = str(entry.get(col, "-")).strip()
        if value not in ["", "-", "nan", "None"]:
            return True

    for elem in ELEMENT_COLUMNS:
        if str(entry.get(elem, "-")).strip() in ["0", "1"]:
            return True

    return False


def make_fingerprint(entry: Dict[str, Any]) -> str:
    """
    生成去重指纹。
    同一个 nanozyme 如果对应不同 substrate / Km / Vmax / Kcat，会保留为不同记录。
    """
    keys = [
        "Nanozyme Name",
        "Classification",
        "Substrate",
        "Optimal pH",
        "Optimal temperature (C)",
        "Km (mM)",
        "Vmax (uM s-1)",
        "Kcat (s-1)",
        "Shape",
        "Size (nm)"
    ]

    return "||".join([str(entry.get(k, "-")).strip().lower() for k in keys])


# ================= 5. 核心推理函数 =================
def extract_nanozyme_data_segmented(md_content: str, target_name: str, doi: str) -> List[Dict[str, Any]]:
    """
    对全文进行分段推理并汇总结果。
    宽松模式：只要有纳米酶名称、分类、组成、形貌、动力学参数等信息，就保留。
    """
    chunks = split_text_into_chunks(md_content)
    all_results = []
    seen_fingerprints = set()

    detected_candidates = extract_nanozyme_candidates(md_content)

    print(f"  🧪 检测到可能的纳米酶名称: {detected_candidates[:10]}")

    candidate_hint = ""
    if detected_candidates:
        candidate_hint = (
            "\n\n### 文档中可能出现的 nanozyme 名称：\n"
            + ", ".join(detected_candidates[:20])
            + "\n请优先检查这些名称是否对应真实 nanozyme。"
        )

    element_keys = "\n".join([f'      "{e}": "1/0/-",' for e in ELEMENT_COLUMNS])

    property_keys = """
      "Metal oxidation state": "-",
      "Shape": "-",
      "Size (nm)": "-",
      "Surface treatment": "-",
      "Dispersion medium": "-",
      "Optimal pH": "-",
      "Optimal temperature (C)": "-",
      "Substrate": "-",
      "Km (mM)": "-",
      "Vmax (uM s-1)": "-",
      "Kcat (s-1)": "-"
"""

    for i, chunk in enumerate(chunks):
        user_prompt = f"""请针对论文 DOI【{doi}】中的纳米酶材料提取数据。
当前参考目标名称为：【{target_name}】。
这是论文的第 {i + 1}/{len(chunks)} 部分。
{candidate_hint}

### 期望 JSON 格式：
[
  {{
    "Nanozyme Name": "{target_name}",
    "Classification": "POD/OXD/CAT/SOD/HYL 或多个用分号连接",
    "elements": {{
{element_keys}
    }},
    "properties": {{
{property_keys}
    }}
  }}
]

### 再次强调：
- 如果当前片段没有动力学参数，也要尽量提取 nanozyme name、classification、元素组成、形貌、尺寸、表面修饰、分散介质等信息。
- 如果当前片段没有任何纳米酶相关信息，请输出 []。
- 没有明确出现的信息填 "-"，不要猜测。
- 多个 nanozyme 或多个 substrate 对应不同动力学参数时，请拆成多条 JSON 记录。

Document Content:
{chunk}
"""

        try:
            resp = CLIENT.chat.completions.create(
                model=MODEL_NAME,
                messages=[
                    {"role": "system", "content": SYSTEM_PROMPT},
                    {"role": "user", "content": user_prompt}
                ],
                temperature=0.0
            )

            content = resp.choices[0].message.content.strip()
            parsed = parse_llm_json(content)

            if not parsed:
                print(f"  ⚠️ 片段 {i + 1} 未解析到有效 JSON")
                continue

            for raw_entry in parsed:
                if not isinstance(raw_entry, dict):
                    continue

                entry = normalize_entry(raw_entry, target_name)

                if not is_useful_entry(entry):
                    continue

                fingerprint = make_fingerprint(entry)

                if fingerprint not in seen_fingerprints:
                    all_results.append(entry)
                    seen_fingerprints.add(fingerprint)

                    print(
                        f"    ✓ 提取到: {entry.get('Nanozyme Name', '-')} | "
                        f"{entry.get('Classification', '-')} | "
                        f"Substrate={entry.get('Substrate', '-')}"
                    )

        except json.JSONDecodeError as e:
            print(f"  ⚠️ 片段 {i + 1} JSON解析失败: {e}")
        except Exception as e:
            print(f"  ⚠️ 片段 {i + 1} 提取失败: {e}")

    return all_results


# ================= 6. 主程序 =================
def main():
    print("🚀 纳米酶全文本分段提取系统启动...")

    # Excel 初始化
    if Path(OUTPUT_EXCEL).exists():
        wb = load_workbook(OUTPUT_EXCEL)
        ws = wb.active

        try:
            df_exist = pd.read_excel(OUTPUT_EXCEL)
            if not df_exist.empty and "DOI" in df_exist.columns:
                existing_dois = set(df_exist["DOI"].astype(str).str.lower())
            else:
                existing_dois = set()
        except Exception:
            existing_dois = set()

    else:
        wb = Workbook()
        ws = wb.active
        ws.append(ALL_COLUMNS)
        existing_dois = set()

    # 读取输入 Excel
    try:
        df_ref = pd.read_excel(INPUT_EXCEL, header=0)

        if df_ref.empty:
            print("⚠️ 输入 Excel 文件为空，将处理所有 DOI 文件夹")
            ref_dict = {}
        else:
            if "DOI" not in df_ref.columns:
                print("❌ 输入 Excel 中必须包含 DOI 列")
                return

            ref_dict = {
                str(row["DOI"]).strip().lower(): row.to_dict()
                for _, row in df_ref.iterrows()
                if pd.notna(row.get("DOI", None))
            }

    except Exception as e:
        print(f"❌ 配置文件读取失败: {e}")
        return

    if not CACHE_DIR.exists():
        print(f"❌ OCR 结果目录不存在: {CACHE_DIR}")
        return

    doi_folders = [d for d in CACHE_DIR.iterdir() if d.is_dir()]

    for folder in tqdm(doi_folders, desc="总体进度"):
        original_doi = folder.name.replace("_", "/")

        if original_doi.lower() in existing_dois:
            continue

        md_files = list(folder.glob("*.md"))

        if not md_files:
            continue

        # 如果一个 DOI 文件夹中有多个 md，则合并读取
        content_parts = []
        for md_file in sorted(md_files):
            try:
                content_parts.append(md_file.read_text(encoding="utf-8"))
            except UnicodeDecodeError:
                content_parts.append(md_file.read_text(encoding="utf-8", errors="ignore"))

        content = "\n\n".join(content_parts)

        info = ref_dict.get(original_doi.lower(), {})

        # 智能目标名称，多级回退
        target_name = pick_first_valid(
            info,
            [
                "Nanozyme Name",
                "Nanozyme",
                "nanozyme",
                "Material",
                "material",
                "Catalyst",
                "catalyst",
                "Samples Name",
                "Sample Name"
            ],
            default=""
        )

        if not target_name:
            candidates = extract_nanozyme_candidates(content)

            if candidates:
                target_name = candidates[0]
            else:
                if USE_DOI_IN_NAME:
                    doi_clean = original_doi.replace("/", "-").replace(".", "_")
                    numbers = re.findall(r"\d+", doi_clean)

                    if numbers:
                        target_name = f"{DEFAULT_NAME_PREFIX}-{numbers[0]}"
                    else:
                        target_name = f"{DEFAULT_NAME_PREFIX}-{doi_clean[:8]}"
                else:
                    target_name = f"{DEFAULT_NAME_PREFIX}-Unknown"

        print(f"\n📄 处理中: {original_doi} (目标: {target_name})")

        data_list = extract_nanozyme_data_segmented(content, target_name, original_doi)

        if not data_list:
            print("  ⚠️ 未发现任何纳米酶有效数据")
            continue

        print(f"  ✅ 成功提取 {len(data_list)} 条记录")

        for item in data_list:
            row_vals = []

            for col in ALL_COLUMNS:
                if col == "DOI":
                    row_vals.append(original_doi)
                elif col == "Year":
                    row_vals.append(info.get("Year", "-"))
                elif col == "Nanozyme Name":
                    row_vals.append(item.get("Nanozyme Name", target_name))
                else:
                    row_vals.append(item.get(col, "-"))

            ws.append(row_vals)

        wb.save(OUTPUT_EXCEL)

    print(f"\n✅ 任务完成！结果已存入: {OUTPUT_EXCEL}")


if __name__ == "__main__":
    main()